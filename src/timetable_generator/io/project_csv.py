"""Project CSV/Excel import & export — FR-017 project parameter ingestion.

CSV/Excel format (UTF-8, optional BOM for CSV):
    项目标识,项目名称,业务线,投入百分比,项目开始时间,项目结束时间

Notes:
    * 投入百分比 is a float ratio in [0, 1] (e.g. 0.3, not 30%).
    * 项目开始/结束时间 use ISO format YYYY-MM-DD; blank falls back to the
      caller-supplied default_start / default_end.
    * `required_job_types` defaults to an empty list (no job-type constraint).
    * `associated_person_ids` defaults to ``["__pending__"]`` to satisfy
      Project's non-empty invariant until the user associates staff via the UI.
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from timetable_generator.models.project import Project

REQUIRED_COLUMNS = [
    "项目标识",
    "项目名称",
    "业务线",
    "投入百分比",
    "项目开始时间",
    "项目结束时间",
]
HEADER_ROW = REQUIRED_COLUMNS
PENDING_PERSON_ID = "__pending__"


def _clean(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def _parse_date_cell(value: str | None) -> date | None:
    raw = _clean(value)
    if not raw:
        return None
    return date.fromisoformat(raw)


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)
    return fieldnames, rows


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    from datetime import datetime

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    header_cols = [("" if c is None else str(c)) for c in header]
    data_rows: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: dict[str, str] = {}
        for col, cell in zip(header_cols, raw, strict=False):
            if col == "":
                continue
            if isinstance(cell, datetime):
                row[col] = cell.date().isoformat()
            elif isinstance(cell, date):
                row[col] = cell.isoformat()
            else:
                row[col] = "" if cell is None else str(cell)
        data_rows.append(row)
    wb.close()
    return header_cols, data_rows


def import_projects_csv(path: Path, default_start: date, default_end: date) -> list[Project]:
    """Import projects from a CSV or Excel file at *path* (6 columns).

    `required_job_types` defaults to an empty list. `associated_person_ids`
    defaults to ``["__pending__"]`` to satisfy Project's non-empty invariant
    until the user associates staff via the UI.
    Dates fall back to *default_start* / *default_end* when blank.
    """
    fieldnames, rows = _read_rows(path)
    missing = [c for c in REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        raise ValueError(f"project file missing required columns: {missing}")

    projects: list[Project] = []
    for row in rows:
        pid = _clean(row.get("项目标识"))
        if not pid:
            continue  # skip blank rows
        name = _clean(row.get("项目名称")) or pid
        business_line = _clean(row.get("业务线"))
        ratio_raw = _clean(row.get("投入百分比"))
        target_ratio = float(ratio_raw) if ratio_raw else 0.0
        start_date = _parse_date_cell(row.get("项目开始时间")) or default_start
        end_date = _parse_date_cell(row.get("项目结束时间")) or default_end
        projects.append(
            Project(
                id=pid,
                name=name,
                start_date=start_date,
                end_date=end_date,
                target_ratio=target_ratio,
                required_job_types=[],
                associated_person_ids=[PENDING_PERSON_ID],
                business_line=business_line,
            )
        )
    return projects


def _format_date(d: date | None) -> str:
    return d.isoformat() if d is not None else ""


def _project_row(p: Project) -> list[str]:
    return [
        p.id,
        p.name,
        p.business_line or "",
        f"{p.target_ratio}",
        _format_date(p.start_date),
        _format_date(p.end_date),
    ]


def export_projects_csv(projects: list[Project], path: Path) -> None:
    """Export *projects* to a CSV or Excel file at *path* (6 columns)."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        _export_xlsx(projects, path)
    else:
        _export_csv(projects, path)


def _export_csv(projects: list[Project], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADER_ROW)
        for p in projects:
            writer.writerow(_project_row(p))


def _export_xlsx(projects: list[Project], path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "项目"
    ws.append(HEADER_ROW)
    for p in projects:
        ws.append(
            [
                p.id,
                p.name,
                p.business_line or "",
                p.target_ratio,
                p.start_date,
                p.end_date,
            ]
        )
    wb.save(path)
