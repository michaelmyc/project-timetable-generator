"""Staff CSV/Excel import & export — FR-017 staff parameter ingestion.

CSV/Excel format (UTF-8, optional BOM for CSV):
    员工,工种,业务线,入职时间,离职时间
Defaults:
    empty 工种 → "研发人员"; empty 业务线 → None;
    empty 入职时间/离职时间 → None (ISO format YYYY-MM-DD when present).
"""

from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from timetable_generator.io.csv_reader import read_csv_rows
from timetable_generator.models.staff_info import DEFAULT_JOB_TYPE, StaffInfo

REQUIRED_COLUMNS = ["员工", "工种", "业务线", "入职时间", "离职时间"]
HEADER_ROW = REQUIRED_COLUMNS


def _clean(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = value.strip()
    return stripped or None


def _parse_date_cell(value: str | None) -> date | None:
    raw = _clean(value)
    if not raw:
        return None
    return date.fromisoformat(raw)


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """Return (header_columns, rows-as-dicts). Dispatches by extension."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    return _read_csv(path)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    return read_csv_rows(path)


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    from datetime import datetime

    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    header_cols = [("" if c is None else str(c)) for c in header]
    data_rows: list[dict[str, str]] = []
    for raw in rows_iter:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        row: dict[str, str] = {}
        for col, cell in zip(header_cols, raw, strict=False):
            if col == "":
                continue
            if isinstance(cell, datetime):
                row[col] = cell.date().isoformat()
            elif isinstance(cell, date):
                row[col] = cell.isoformat()
            else:
                row[col] = "" if cell is None else str(cell)
        data_rows.append(row)
    wb.close()
    return header_cols, data_rows


def import_staff_csv(path: Path) -> list[StaffInfo]:
    """Import staff profiles from a CSV or Excel file at *path*.

    The first row must be a header containing the columns
    员工, 工种, 业务线, 入职时间, 离职时间 (in any order). Blank data rows are
    skipped.
    """
    fieldnames, rows = _read_rows(path)
    missing = [c for c in REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        raise ValueError(f"staff file missing required columns: {missing}")

    staff: list[StaffInfo] = []
    for row in rows:
        name = _clean(row.get("员工"))
        if not name:
            continue  # skip blank rows
        job_type = _clean(row.get("工种")) or DEFAULT_JOB_TYPE
        business_line = _clean(row.get("业务线"))
        onboard_date = _parse_date_cell(row.get("入职时间"))
        leave_date = _parse_date_cell(row.get("离职时间"))
        staff.append(
            StaffInfo(
                name=name,
                job_type=job_type,
                business_line=business_line,
                onboard_date=onboard_date,
                leave_date=leave_date,
            )
        )
    return staff


def _format_date(d: date | None) -> str:
    return d.isoformat() if d is not None else ""


def _staff_row(s: StaffInfo) -> list[str]:
    return [
        s.name,
        s.job_type,
        s.business_line or "",
        _format_date(s.onboard_date),
        _format_date(s.leave_date),
    ]


def export_staff_csv(staff: list[StaffInfo], path: Path) -> None:
    """Export *staff* to a CSV or Excel file at *path* (5 columns)."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        _export_xlsx(staff, path)
    else:
        _export_csv(staff, path)


def _export_csv(staff: list[StaffInfo], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADER_ROW)
        for s in staff:
            writer.writerow(_staff_row(s))


def _export_xlsx(staff: list[StaffInfo], path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "员工"
    ws.append(HEADER_ROW)
    for s in staff:
        ws.append(
            [
                s.name,
                s.job_type,
                s.business_line or "",
                s.onboard_date,
                s.leave_date,
            ]
        )
    wb.save(path)
